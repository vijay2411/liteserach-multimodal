"""XLSX extractor — one segment per sheet, rows as TSV-ish text, via openpyxl."""
from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from semanticsd.extractors.base import Extractor, ExtractedDoc, ExtractedSegment
from semanticsd.extractors.registry import register


@register
class XlsxExtractor(Extractor):
    file_type = "xlsx"
    extensions = (".xlsx", ".xlsm")

    def extract(self, path: Path) -> ExtractedDoc:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        segments: list[ExtractedSegment] = []
        cursor = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                rows.append("\t".join(cells))
            text = "\n".join(rows).strip()
            if not text:
                continue
            seg_bytes = len(text.encode("utf-8"))
            segments.append(ExtractedSegment(
                text=text,
                byte_start=cursor,
                byte_end=cursor + seg_bytes,
                metadata={"sheet": sheet_name},
            ))
            cursor += seg_bytes + 1
        wb.close()
        return ExtractedDoc(
            path=str(path),
            file_type=self.file_type,
            segments=segments,
        )
